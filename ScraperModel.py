import os
import time
import usaddress
from openpyxl import load_workbook
import undetected_chromedriver as uc
import bs4
import re

class DataScraper:
    def __init__(self, chromedriver_path="chromedriver.exe", xlsx_path="data.xlsx"):
        self.chromedriver_path = chromedriver_path
        self.xlsx_path = xlsx_path
        self.curr_script_path = os.path.realpath(os.path.dirname(__file__))
        self.profile_path = self.curr_script_path + "\\profile"
        self.first_name_col = 'A'
        self.last_name_col = 'B'
        self.address_col = 'F'
        self.phones_cols = ['L', 'M', 'N', 'O', 'P']
        self.controller = None

    def setController(self, controller):
        self.controller = controller

    def open_chrome_with_profile(self):
        options = uc.ChromeOptions()
        options.add_argument("--user-data-dir=" + self.profile_path)
        driver = uc.Chrome(driver_executable_path=self.chromedriver_path, options=options)
        return driver

    def set_xlsx_file(self, filename):
        self.xlsx_path = filename.strip()

    def open_xlsx_file(self):
        print(self.xlsx_path)
        wb = load_workbook(self.xlsx_path)
        ws = wb.active
        return wb, ws

    def write_phones_to_xlsx_file(self, wb, ws, phones, row):
        phone_length = min(len(phones), 5)
        for i in range(phone_length):
            ws[self.phones_cols[i] + str(row)].value = phones[i]
        wb.save(self.xlsx_path)

    def savePhonesToDatabase(self, clientID, phones):
        phone_length = min(len(phones), 5)
        for i in range(phone_length):
            self.controller.addPhone(self.controller.retrievePhones(1)[0][0] + 1, clientID, phones[i])

    def extract_phones_from_page(self, page_source):
        phones = []
        try:
            soup = bs4.BeautifulSoup(page_source, "html.parser")
            a_tags = soup.find_all("a", title=lambda x: x and "Search people with phone number" in x)
            for a_tag in a_tags:
                phones.append(a_tag.text.strip())
        except Exception as e:
            print(str(e))
        return phones

    def extract_address_from_page(self, page_source):
        address = []
        try:
            soup = bs4.BeautifulSoup(page_source, "html.parser")
            a_tags = soup.find_all("a", title=lambda  x: x and "Search people living at" in x)
            for a_tag in a_tags:
                tempAddress = a_tag.text.strip()
                address.append(re.sub(r'\n', ' ', tempAddress))
        except Exception as e:
            print(str(e))

        address = self.parseAddress(address[0])
        return address

    def parseAddress(self, address):
        parsed = usaddress.tag(address)

        if 'StreetName' in parsed[0]:
            streetAddress = ' '.join([parsed[0].get(part, '') for part in ['AddressNumber', 'StreetNamePreDirectional', 'StreetName', 'StreetNamePostType']])
            city = parsed[0].get('PlaceName', '')
            state = parsed[0].get('StateName', '')
            zipcode = parsed[0].get('ZipCode', '')
            return {'streetAddress': streetAddress.strip(), 'city': city, 'state': state, 'zipcode': zipcode}
        return None

    def extract_age_from_page(self, page_source):
        try:
            soup = bs4.BeautifulSoup(page_source, "html.parser")
            div_tag = soup.find('div', class_='card-block')

            age_tag = div_tag.find('h3', text='Age:')

            if age_tag:
                age_text = age_tag.find_next_sibling(text=True).strip()
                #If current client searched is dead, make age 0
                if age_text.startswith("Deceased"):
                    age_text = 0
                return age_text
            else:
                return None

        except Exception as e:
            print(str(e))
            return None

    def run(self):
        driver = self.open_chrome_with_profile()
        driver.get("https://www.fastpeoplesearch.com/")
        if "Access Denied" in driver.page_source:
            print("Access Denied, please enable VPN")
            time.sleep(60)
            driver.get("https://www.fastpeoplesearch.com/")
            if "Access Denied" in driver.page_source:
                return 1

        wb, ws = self.open_xlsx_file()
        for row in range(2, ws.max_row + 1):
            try:
                first_name = ws[self.first_name_col + str(row)].value
                last_name = ws[self.last_name_col + str(row)].value
                address = ws[self.address_col + str(row)].value

                if (first_name is None and last_name is None) or address is None:
                    continue

                search_url = f"https://www.fastpeoplesearch.com/name/{first_name.replace(' ', '-')}-{last_name.replace(' ', '-')}_{str(address).replace(' ', '-')}"
                driver.get(search_url)
                phones = self.extract_phones_from_page(driver.page_source)
                address = self.extract_address_from_page(driver.page_source)
                age = self.extract_age_from_page(driver.page_source)
                print(age)
                if phones and address and age:
                    self.write_phones_to_xlsx_file(wb, ws, phones, row)
                    tempClientID = self.controller.retrieveClients(1)[0][0]
                    self.controller.addClient(tempClientID + 1, self.controller.currentUserID, first_name, last_name, "Life", int(age))
                    self.savePhonesToDatabase(tempClientID, phones)
                    #ADD CITIES
                    self.controller.addAddress(self.controller.retrieveAddresses(1)[0][0] + 1, tempClientID, address.get('streetAddress'), address.get('state'), address.get('zipcode'))

                else:
                    print(f"No phones found for {first_name} {last_name}")
                time.sleep(1)
            except Exception as e:
                print(str(e))
                continue

        wb.close()
        driver.close()